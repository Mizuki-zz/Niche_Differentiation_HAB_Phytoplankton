import pandas as pd
import numpy as np
from scipy import stats
from scikit_posthocs import posthoc_dunn
import openpyxl
from openpyxl.styles import Font


def p_value_to_stars(p_value):
    """Convert p-value to asterisk notation"""
    if p_value < 0.001:
        return "***"
    elif p_value < 0.01:
        return "**"
    elif p_value < 0.05:
        return "*"
    else:
        return "ns"


def epsilon_squared(H, n):
    """
    Calculate effect size epsilon-squared (ε²) for Kruskal-Wallis test
    Formula: ε² = H / (n - 1)
    where H is the Kruskal-Wallis statistic and n is the total sample size
    """
    return H / (n - 1)


def effect_size_interpretation(epsilon_sq):
    """
    Interpret epsilon-squared effect size based on empirical thresholds
    Small: < 0.01
    Medium: 0.01 – 0.08
    Large: ≥ 0.08
    (Note: thresholds vary slightly by field; common standards are used here)
    """
    if epsilon_sq < 0.01:
        return "small"
    elif epsilon_sq < 0.08:
        return "medium"
    else:
        return "large"


# Read Excel file
file_path = "ND_FD.xlsx"
sheets = ['d', 'e', 'f', 'g1', 'g2', 'h1', 'h2']  # List of sheets that output summary results

# Create a new Excel workbook to save results
result_wb = openpyxl.Workbook()

# Remove the default worksheet
if 'Sheet' in result_wb.sheetnames:
    result_wb.remove(result_wb['Sheet'])

# Create a result sheet for each original sheet
for sheet_name in sheets:
    # Read data
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    # Get list of seasons
    seasons = sorted(df['season'].unique())

    # Create result worksheet
    result_ws = result_wb.create_sheet(title=f"{sheet_name}_results")

    # ---------- Kruskal-Wallis test results ----------
    result_ws['A1'] = "Kruskal-Wallis Test Results"
    result_ws['A1'].font = Font(bold=True)

    result_ws['A2'] = "H statistic"
    result_ws['B2'] = "P-value"
    result_ws['C2'] = "Significance"
    result_ws['D2'] = "Effect size (ε²)"      # Added
    result_ws['E2'] = "Effect magnitude"       # Added

    # Perform Kruskal-Wallis test
    groups = [df[df['season'] == season]['index'].values for season in seasons]
    h_stat, p_val = stats.kruskal(*groups)

    # Calculate total sample size n
    n_total = sum(len(g) for g in groups)
    # Calculate effect size ε²
    eps2 = epsilon_squared(h_stat, n_total)
    # Qualitative effect size interpretation
    eff_size = effect_size_interpretation(eps2)

    # Write Kruskal-Wallis results (new columns for effect size and interpretation)
    result_ws['A3'] = h_stat
    result_ws['B3'] = p_val
    result_ws['C3'] = p_value_to_stars(p_val)
    result_ws['D3'] = eps2
    result_ws['E3'] = eff_size

    # ---------- Dunn's post-hoc test results ----------
    result_ws['A5'] = "Dunn's Post-hoc Test Results (Bonferroni correction)"
    result_ws['A5'].font = Font(bold=True)

    # Perform Dunn's test
    dunn_result = posthoc_dunn(groups, p_adjust='bonferroni')

    # Write row and column labels
    for i, season in enumerate(seasons):
        result_ws.cell(row=6, column=5 + i).value = season
        result_ws.cell(row=7 + i, column=4).value = season

    # Write Dunn's test results
    for i in range(len(seasons)):
        for j in range(len(seasons)):
            if i != j:
                p_val_dunn = dunn_result.iloc[i, j]
                result_ws.cell(row=7 + i, column=5 + j).value = f"{p_val_dunn:.4f} ({p_value_to_stars(p_val_dunn)})"
            else:
                result_ws.cell(row=7 + i, column=5 + j).value = "-"

    # ---------- Sample size information ----------
    result_ws['A10'] = "Sample sizes per group"  # Row number adjusted to avoid conflict
    result_ws['A10'].font = Font(bold=True)

    for i, season in enumerate(seasons):
        result_ws.cell(row=11, column=1 + i).value = season
        result_ws.cell(row=12, column=1 + i).value = len(df[df['season'] == season])

    # Adjust column widths
    for column in result_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 20)
        result_ws.column_dimensions[column_letter].width = adjusted_width

# Save the result
result_wb.save("results.xlsx")
print("Analysis complete! Results saved to 'results.xlsx'")